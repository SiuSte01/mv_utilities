#!/opt/local/marketview/conda/envs/envs/dev/bin/python

from dateutil.parser import parse as dtParse
import logging
import argparse
import os, sys
import subprocess
import warnings
import json
import yaml
import pandas as pd
import numpy as np
import cx_Oracle as cxo
from contextlib import closing
import itertools, collections
from tqdm import tqdm
import time
import re
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import openpyxl.styles as xlStyles
import aggr
from mvUtil import miscfunctions as mf

logger = logging.getLogger(__name__)
oraInst = None
oraUser = None
oraPass = None
memoryHash = {}
setVars = {}
superProjDir = ""

# oraUser = "claims_aggr"
# oraPass = "Hydr0gen2014"
# oraInst = "PLDWH2DBR"
# sql = "select * from profiledata.organizations_view t where t.HMS_POID = 'POZ2242B57' and vintage_num = 20210818"
#
# selDF = mf.getOracleSql(oraInst,oraUser,oraPass,sql)
# print(selDF)
# exit()


def main():
    # tqdmLoop()
    parser = argparse.ArgumentParser(description="Read in a config file.")
    parser.add_argument("-config", "--config", required=True)

    args = parser.parse_args()
    # print(args)
    # print(args.config)

    configPath = os.path.abspath(args.config)
    projDir = os.path.dirname(configPath)
    global superProjDir, setVars, memoryHash, oraInst, oraUser, oraPass
    superProjDir = os.path.dirname(projDir)
    try:
        os.makedirs(superProjDir + "/milestones")
    except:
        pass

    setVars = mf.createSettingDict(args.config)
    # print(yaml.dump(setVars))
    oraInst = setVars["INSTANCE"][0]
    oraUser = setVars["USERNAME"][0]
    oraPass = setVars["PASSWORD"][0] if setVars["PASSWORD"] else "Hydr0gen2014"
    jobName = setVars["JOB_NAME"][0]
    vint = mf.normalizeDate(setVars["VINTAGE"][0])
    repType = setVars["REPORT_TYPE"][0]
    dosWin = setVars["DOS_WINDOW"][0] if setVars["DOS_WINDOW"] else ""
    procCodesFile = setVars["PROCEDURE_CODES_FILE"][0]
    adjustFile = setVars["ADJUSTMENTS_FILE"][0]
    # print('-'.join([oraInst,oraUser,oraPass,jobName,vint,repType,dosWin,procCodesFile,adjustFile]))

    # run either distribution report, adjustment code analysis (avanced remit), or both
    if repType.lower() == "dist":
        buildDist()
    elif repType.lower() == "aca":
        buildAca()
    elif repType.lower() == "both":
        buildDist()
        buildAca()
    else:
        exit("Unrecognized REPORT_TYPE: " + repType)


def tqdmLoop():
    for i in tqdm(range(0, 100), desc="tqdm sample"):
        sleep(0.5)


def buildDist():
    distFile = "priPlusDistributionReport.tab"
    print("Building Distribution Report: " + distFile + "...", end="", flush=True)
    global superProjDir, setVars, memoryHash, oraInst, oraUser, oraPass
    jobId = -1
    try:
        jobId = setVars["JOB_ID"][0]
    except:
        exit("Cannot determine JOB_ID\n")
    ppdcTable = "PRIPLUSDISTCLAIMS" + jobId
    pppTable = "PRIPLUSPIVOT" + jobId
    # determine if table already exists and needs to be dropped
    for x in (ppdcTable, pppTable):
        tableExistsSql = (
            """select count(*) as Table_Exists
                                    from user_tables
                                    where table_name = '"""
            + x
            + "'"
            ""
        )
        teRes = mf.getOracleSql(oraInst, oraUser, oraPass, tableExistsSql)[
            "TABLE_EXISTS"
        ].iloc[0]
        if teRes:
            mf.dropTable(oraInst, oraUser, oraPass, x)
    fltrSql = (
        """create table """
        + ppdcTable
        + """ as (select distinct remit_claim_uid,
		svc_line_seq,
		procedure_code,
		svc_adjust_grp1,
		svc_adjust_grp2,
		svc_adjust_grp3,
		svc_adjust_grp4,
		svc_adjust_grp5,
		svc_adjust_grp6
		from pri_plus_claims where job_id = """
        + jobId
        + """)"""
    )
    mf.getOracleSql(oraInst, oraUser, oraPass, fltrSql)
    pivotSql = (
        """create table """
        + pppTable
        + """ as (
        select adj.*,
                case
                  when adj.adjustGrp = 0
                    then ''
                  when exists (select * from claimswh.adjustment_group_members am where adj.adjustgrp = am.adjust_group_id)
                    then (
                        select distinct(adjust_reason_code) from (
                            select cast(am.adjust_reason_id as varchar(255)),
                                    ar.adjust_reason_code from claimswh.adjustment_group_members am
                            left join claimswh.adjustment_reasons ar on am.adjust_reason_id = ar.adjust_reason_id
                            where adj.adjustgrp = am.adjust_group_id)
                    )
                  else 'UNKWN'
                  end as adjust_reason_code
                  from (
                        select remit_claim_uid,svc_line_seq,procedure_code,svc_adjust_grp1 as adjustGrp from """
        + ppdcTable
        + """ grp1
                            union
                        select remit_claim_uid,svc_line_seq,procedure_code,svc_adjust_grp2 as adjustGrp from """
        + ppdcTable
        + """ grp2
                            union
                        select remit_claim_uid,svc_line_seq,procedure_code,svc_adjust_grp3 as adjustGrp from """
        + ppdcTable
        + """ grp3
                            union
                        select remit_claim_uid,svc_line_seq,procedure_code,svc_adjust_grp4 as adjustGrp from """
        + ppdcTable
        + """ grp4
                            union
                        select remit_claim_uid,svc_line_seq,procedure_code,svc_adjust_grp5 as adjustGrp from """
        + ppdcTable
        + """ grp5
                            union
                        select remit_claim_uid,svc_line_seq,procedure_code,svc_adjust_grp6 as adjustGrp from """
        + ppdcTable
        + """ grp6
                  ) adj
        )"""
    )
    mf.getOracleSql(oraInst, oraUser, oraPass, pivotSql)
    finalSql = (
        """select arCounts.procedure_code,
                        arCounts.adjust_reason_code,
                        round(arCounts.adjust_reason_cnt/denom.denom,4) as pct_proc_code,
                        ar.reason
                        from (
                            select procedure_code,
                                    adjust_reason_code,
                                    count(adjust_reason_code) as adjust_reason_cnt
                                    from (
                                        select distinct piv.remit_claim_uid,piv.svc_line_seq,piv.procedure_code,999999999 as adjustGrp,piv.adjust_reason_code
                                                from """
        + pppTable
        + """ piv where piv.adjust_reason_code = 'UNKWN'
                                            union
                                        select * from """
        + pppTable
        + """ piv where piv.adjust_reason_code != 'UNKWN'
                                        )
                        group by procedure_code,adjust_reason_code) arCounts
                    left join
                    (
                        select procedure_code,
                                count(procedure_code) as denom
                                from """
        + ppdcTable
        + """ group by procedure_code
                    ) denom on arCounts.procedure_code = denom.procedure_code
                    left join (select distinct adjust_reason_code,description as reason from claimswh.adjustment_reasons) ar on arCounts.adjust_reason_code = ar.adjust_reason_code
                    where arCounts.adjust_reason_code != '0'
                    order by procedure_code,
                                to_number(regexp_substr(adjust_reason_code,'^[0-9]+')),
                                to_number(regexp_substr(adjust_reason_code,'[0-9]+$')),
                                adjust_reason_code"""
    )
    ddf = mf.getOracleSql(oraInst, oraUser, oraPass, finalSql)
    ddf.to_csv(
        superProjDir + "/milestones/priPlusDistributionReport.tab",
        sep="\t",
        index=False,
    )
    for table in (ppdcTable, pppTable):
        mf.dropTable(oraInst, oraUser, oraPass, table)
    print("done")


def buildAca():
    print("Running Adjustment Code Analysis...")
    global superProjDir
    runFlats = 0
    for flatFile in ("Report", "Report_POS", "Report_Mod", "Report_POSMod"):
        if not os.path.isfile(superProjDir + "/milestones/" + flatFile + ".tab"):
            runFlats = 1
    runXL = 1
    if runFlats:
        runXL = buildReportFiles()
    if runXL:
        buildXLSheet()

    print("done")


def buildReportFiles():
    global superProjDir, setVars, memoryHash, oraInst, oraUser, oraPass
    jobId = -1
    try:
        jobId = setVars["JOB_ID"][0]
    except:
        exit("Cannot determine JOB_ID\n")
    vint = mf.normalizeDate(setVars["VINTAGE"][0], yyyymmdd="Y")
    # impute units
    memoryHash["PROC_PAYER_HASH"] = {}
    memoryHash["PROC_UNITS_HASH"] = {}
    memoryHash["CLAIM_ADJUSTMENT_AMOUNTS"] = {}
    adjustFile = setVars["ADJUSTMENTS_FILE"][0]
    adf = pd.read_csv(adjustFile, sep="\t")
    adf["ADJ_CODE"] = adf["ADJ_CODE"].astype(str)
    # limiting pri_plus_flter_claims to only rows where at least one of the adjust reason codes falls in
    # the user's initial list of adjustment reasons
    userAdjSet = set(adf["ADJ_CODE"].values)
    # the adjustment reasons that are actually found in pri_plus_flter_claims
    usedAdjSet = set()
    # print(useradjSet)
    # print(adf)

    #table definitions
    claimsTab = "pri_plus_claims"
    adjTab = "pri_plus_adjust_codes"
    adjTup = "('" + "','".join(userAdjSet) + "')"
    ppacTable = "PRIPLUSACTCLAIMS" + jobId
    piTable = "PRIPLUSPIIDLIST" + jobId
    poTable = "PRIPLUSPOIDLIST" + jobId
    ppTable = "PRIPLUSPROCPAYER" + jobId
    ippTable = "PRIPLUSIMPPROCPAYER" + jobId
    cppTable = "PRIPLUSCLEANPROCPAYER" + jobId
    ppaccTable = "PRIPLUSACTCLAIMSCLEAN" + jobId
    caaTable = "PRIPLUSCLAIMADJAMTS" + jobId
    prtTable = "PRIPLUSPATIENTRESPTOTAL" + jobId
    dedTable = "PRIPLUSDEDUCT" + jobId
    coinsTable = "PRIPLUSCOINS" + jobId
    copayTable = "PRIPLUSCOPAY" + jobId
    cv1Table = "PRIPLUSCV1" + jobId
    cv2Table = "PRIPLUSCV2" + jobId
    caa2Table = "PRIPLUSCLAIMADJAMTS2" + jobId
    cv3Table = "PRIPLUSCV3" + jobId
    # determine if table already exists and needs to be dropped
    for x in (ppacTable,piTable,poTable,ppTable,ippTable,cppTable,
    ppaccTable,caaTable,prtTable,dedTable,coinsTable,copayTable,
    cv1Table,cv2Table,caa2Table,cv3Table):
        tableExistsSql = (
            """select count(*) as Table_Exists
                                    from user_tables
                                    where table_name = '"""
            + x
            + "'"
            ""
        )
        teRes = mf.getOracleSql(oraInst, oraUser, oraPass, tableExistsSql)[
            "TABLE_EXISTS"
        ].iloc[0]
        if teRes:
            mf.dropTable(oraInst, oraUser, oraPass, x)
    procPayerACSql = getSql("procPayerActivityClaims",{"ppac":ppacTable,"claims":claimsTab,"adj":adjTab,"jobId":jobId,"adjTup":adjTup})
    #print(procPayerACSql)
    mf.getOracleSql(oraInst, oraUser, oraPass, procPayerACSql)
    # create header files if there is no data
    tableEmptySql = (
        """select count(*) as Table_Empty
                            from """
        + ppacTable
    )
    teRes = mf.getOracleSql(oraInst, oraUser, oraPass, tableEmptySql)[
        "TABLE_EMPTY"
    ].iloc[0]
    if teRes == 0:
        print("No data found, nothing to build")
        mf.dropTable(oraInst, oraUser, oraPass, ppacTable)
        return 0
    #print(adf.head(3))
    payPiSql = getSql("payPi",{"pi":piTable,"ppac":ppacTable,"jobId":jobId})
    payPoSql = getSql("payPo",{"po":poTable,"ppac":ppacTable,"jobId":jobId})
    mf.getOracleSql(oraInst, oraUser, oraPass, payPiSql)
    mf.getOracleSql(oraInst, oraUser, oraPass, payPoSql)
    #print(payPiSql)
    #print(payPoSql)
    acaStartTime = time.time()
    print("\tBuilding ACA Data...", end="", flush=True)
    procPayerSql = getSql("procPayer",{"pp":ppTable,"ppac":ppacTable})
    #print(procPayerSql)
    mf.getOracleSql(oraInst,oraUser,oraPass,procPayerSql)
    #procedure_code payer_name combinations where impute_units = 0, for calculating imputation values
    impProcPayerSql = getSql("impProcPayer",{"ipp":ippTable,"pp":ppTable})
    #print(impProcPayerSql)
    mf.getOracleSql(oraInst,oraUser,oraPass,impProcPayerSql)
    #clean_units table for proc/payer pairings
    cleanProcPayerSql = getSql("cleanProcPayer",{"cpp":cppTable,"pp":ppTable,"ipp":ippTable})
    #print(cleanProcPayerSql)
    mf.getOracleSql(oraInst,oraUser,oraPass,cleanProcPayerSql)
    #ppac table with left joined clean_units and dropped data where svc_prcsd_units = 0 and not imputed
    procPayerACCSql = getSql("procPayerActivityClaimsClean",{"ppacc":ppaccTable,"ppac":ppacTable,"ipp":ippTable,"cpp":cppTable})
    #print(procPayerACCSql)
    mf.getOracleSql(oraInst, oraUser, oraPass, procPayerACCSql)
    #caa table for adjustment amount aggregate for remid_claim_service_id/adjust_reason_code pairings where arc is in adjustments.tab
    claimAdjAmtsSql = getSql("claimAdjAmts",{"caa":caaTable,"ppacc":ppaccTable})
    #print(claimAdjAmtsSql)
    mf.getOracleSql(oraInst, oraUser, oraPass, claimAdjAmtsSql)
    #prt table. basically the same data as the caa table, but limited to amounts where adj_reason is 1,2,3 or group_code is 'PR' (mapping claim_id to patientRespTotal)
    patientRespTotalSql = getSql("patientRespTotal",{"prt":prtTable,"ppacc":ppaccTable})
    #print(patientRespTotalSql)
    mf.getOracleSql(oraInst, oraUser, oraPass, patientRespTotalSql)
    #deduct,coins,copay tables. same as prt but for adj_reason = 1,2, and 3, respectively
    deductSql = getSql("cleanCost",{"cln":dedTable,"ppacc":ppaccTable,"ar":"1"})
    coinsSql = getSql("cleanCost",{"cln":coinsTable,"ppacc":ppaccTable,"ar":"2"})
    copaySql = getSql("cleanCost",{"cln":copayTable,"ppacc":ppaccTable,"ar":"3"})
    for sql in (deductSql,coinsSql,copaySql):
        #print(sql)
        mf.getOracleSql(oraInst, oraUser, oraPass, sql)
    #cv1 table for adding clean values except clean_discount and clean_unit values
    cv1Sql = getSql("cleanVals1",{"cv1":cv1Table,"ppacc":ppaccTable,"prt":prtTable,"ded":dedTable,"coins":coinsTable,"copay":copayTable})
    #print(cv1Sql)
    mf.getOracleSql(oraInst, oraUser, oraPass, cv1Sql)
    #cv2 table for adding ded, coins, and copay flag fields and remaining clean_fields (adjamtxunitperchrg fields to be added later to caaTable)
    cv2Sql = getSql("cleanVals2",{"cv2":cv2Table,"cv1":cv1Table,"ipp":ippTable})
    #print(cv2Sql)
    mf.getOracleSql(oraInst, oraUser, oraPass, cv2Sql)
    #caa2 table, which is caa with a left join for adjamtxunitperchrg
    caa2Sql = getSql("claimAdjAmts2",{"caa2":caa2Table,"caa":caaTable,"cv2":cv2Table})
    #print(caa2Sql)
    mf.getOracleSql(oraInst, oraUser, oraPass, caa2Sql)
    #cv3 table, final table with adjustment x amounts added
    cv3Sql = getSql("cleanVals3",{"cv3":cv3Table,"adjArr":adf["ADJ_CODE"].unique(),"cv2":cv2Table,"caa2":caa2Table})
    #print(cv3Sql)
    mf.getOracleSql(oraInst, oraUser, oraPass, cv3Sql)
   
    acaEndTime = time.time()
    acaHours, acaRem = divmod(acaEndTime - acaStartTime, 3600)
    acaMinutes, acaSeconds = divmod(acaRem, 60)
    print(
        "done. Run Time: {:0>2}:{:0>2}:{:05.2f}".format(
            int(acaHours), int(acaMinutes), acaSeconds
        )
    )
    
    flatStartTime = time.time()
    print("\tGenerating flat files...", end="", flush=True)
    #generate sql tables for 4 flat files, write them directly to file, call it a day.
    #report
    reportSql = getSql("acaReport",{"levels":[],"adjArr":sorted(adf["ADJ_CODE"].unique(),key=mf.naturalSort),"pi":piTable,"po":poTable,"cv3":cv3Table})
    #print(reportSql)
    reportDF = mf.getOracleSql(oraInst, oraUser, oraPass, reportSql)
    #print(reportDF.head())
    reportDF.to_csv(
        superProjDir + "/milestones/Report.tab",
        sep="\t",
        index=False,
    )
    #reportMod
    reportModSql = getSql("acaReport",{"levels":["MOD_CODES"],"adjArr":sorted(adf["ADJ_CODE"].unique(),key=mf.naturalSort),"pi":piTable,"po":poTable,"cv3":cv3Table})
    #print(reportModSql)
    reportModDF = mf.getOracleSql(oraInst, oraUser, oraPass, reportModSql)
    #print(reportModDF.head())
    reportModDF.to_csv(
        superProjDir + "/milestones/Report_Mod.tab",
        sep="\t",
        index=False,
    )
    #reportPos
    reportPosSql = getSql("acaReport",{"levels":["SETTING"],"adjArr":sorted(adf["ADJ_CODE"].unique(),key=mf.naturalSort),"pi":piTable,"po":poTable,"cv3":cv3Table})
    #print(reportPosSql)
    reportPosDF = mf.getOracleSql(oraInst, oraUser, oraPass, reportPosSql)
    #print(reportPosDF.head())
    reportPosDF.to_csv(
        superProjDir + "/milestones/Report_POS.tab",
        sep="\t",
        index=False,
    )
    #reportPosMod
    reportPosModSql = getSql("acaReport",{"levels":["SETTING","MOD_CODES"],"adjArr":sorted(adf["ADJ_CODE"].unique(),key=mf.naturalSort),"pi":piTable,"po":poTable,"cv3":cv3Table})
    #print(reportPosModSql)
    reportPosModDF = mf.getOracleSql(oraInst, oraUser, oraPass, reportPosModSql)
    #print(reportPosModDF.head())
    reportPosModDF.to_csv(
        superProjDir + "/milestones/Report_POSMod.tab",
        sep="\t",
        index=False,
    )
    flatEndTime = time.time()
    flatHours, flatRem = divmod(flatEndTime - flatStartTime, 3600)
    flatMinutes, flatSeconds = divmod(flatRem, 60)
    print(
        "done. Run Time: {:0>2}:{:0>2}:{:05.2f}".format(
            int(flatHours), int(flatMinutes), flatSeconds
        )
    )
    print("done")
    for x in (ppacTable,piTable,poTable,ppTable,ippTable,cppTable,
    ppaccTable,caaTable,prtTable,dedTable,coinsTable,copayTable,
    cv1Table,cv2Table,caa2Table,cv3Table):
        tableExistsSql = (
            """select count(*) as Table_Exists
                                    from user_tables
                                    where table_name = '"""
            + x
            + "'"
            ""
        )
        teRes = mf.getOracleSql(oraInst, oraUser, oraPass, tableExistsSql)[
            "TABLE_EXISTS"
        ].iloc[0]
        if teRes:
            mf.dropTable(oraInst, oraUser, oraPass, x)
    return 1


def buildXLSheet():
    print("\tBuilding finalDeliverable.xlsx...", end="", flush=True)
    global superProjDir, setVars, memoryHash, oraInst, oraUser, oraPass
    jobId = -1
    try:
        jobId = setVars["JOB_ID"][0]
    except:
        exit("Cannot determine JOB_ID\n")
    overviewSql = (
        "select distinct adjust_reason_code,reason from pri_plus_adjust_codes where job_id = "
        + jobId
        + " order by adjust_reason_code"
    )
    con = cxo.connect(oraUser, oraPass, oraInst)
    oDF = pd.read_sql(overviewSql, con=con)
    adjustFile = setVars["ADJUSTMENTS_FILE"][0]
    adf = pd.read_csv(adjustFile, sep="\t")
    adf["ADJ_CODE"] = adf["ADJ_CODE"].astype(str)
    #print(oDF)
    #print(adf)
    xlout = superProjDir + "/milestones/finalDeliverable.xlsx"
    repFile = superProjDir + "/milestones/Report.tab"
    modFile = superProjDir + "/milestones/Report_Mod.tab"
    posFile = superProjDir + "/milestones/Report_POS.tab"
    pmFile = superProjDir + "/milestones/Report_POSMod.tab"
    wb = xl.load_workbook(
        aggr.__path__._path[0] + "/pri/priPlus/PRI_Plus_Workbook_Template.xlsx"
    )
    ws0 = wb["Overview"]
    ws1 = wb.create_sheet("Overall", 1)
    ws2 = wb.create_sheet("Modifier", 2)
    ws3 = wb.create_sheet("Setting", 3)
    ws4 = wb.create_sheet("Setting + Modifier", 4)
    wsMap = {ws1: repFile, ws2: modFile, ws3: posFile, ws4: pmFile}
    thinBorder = xlStyles.borders.Border(
        left=xlStyles.borders.Side(style="thin"),
        right=xlStyles.borders.Side(style="thin"),
        top=xlStyles.borders.Side(style="thin"),
        bottom=xlStyles.borders.Side(style="thin"),
    )
    # populate Overview sheet with ARC info
    posCounter = 0
    for i, row in oDF.iterrows():
        arc = row["ADJUST_REASON_CODE"]
        if arc not in adf["ADJ_CODE"].unique():
            continue
        desc = row["REASON"]
        arCell = ws0.cell(18 + posCounter, 2)
        deCell = ws0.cell(18 + posCounter, 3)
        arCell.value = arc
        deCell.value = desc
        ws0.merge_cells(start_row=18 + posCounter, start_column=3, end_row=18 + posCounter, end_column=12)
        arCell.border = thinBorder
        # border merged cells C17
        styleRange(ws0, "C" + str(18 + posCounter) + ":L" + str(18 + posCounter), border=thinBorder)
        posCounter += 1
    for ws in wsMap.keys():
        inFile = wsMap[ws]
        inDF = pd.read_csv(inFile, sep="\t", dtype=str, keep_default_na=False)
        avgUnitsPos = inDF.columns.get_loc("AVG_UNITS")
        lineCountPos = inDF.columns.get_loc("LINE_COUNT")
        for field in ["AVG_UNITS","LINE_COUNT"]:
            inDF[field] = pd.to_numeric(inDF[field],errors='coerce')
        dollarFields = []
        pctFields = []
        for field in inDF.columns:
            if re.search("(AMTUNIT|COSTUNIT)",field):
                inDF[field] = pd.to_numeric(inDF[field],errors='coerce')
                dollarFields.append(inDF.columns.get_loc(field))
                #print(inDF[field].head)
            if re.search("(ADJAMT|PCT)",field):
                inDF[field] = pd.to_numeric(inDF[field],errors='coerce')
                pctFields.append(inDF.columns.get_loc(field))
        #print(inDF.columns)
        #print(dollarFields)
        #print(inDF.head(3))
        #exit()
        #AMTUNIT or COSTUNIT
        # print(inDF["AVG_UNITS"].head(10))
        # inDF["AVG_UNITS"] = inDF["AVG_UNITS"].astype(float)
        # print(inDF["AVG_UNITS"].head(10))
        # exit()
        colWidths = []
        # print(ws,wsMap[ws])
        # print(inDF.head())
        rowNum = 1
        cellCoord = get_column_letter(len(inDF.columns) + 1) + str(len(inDF) + 1)
        for r in dataframe_to_rows(inDF, index=False, header=True):
            rowNum += 1
            for i in range(len(r)):
                if len(colWidths) > i:
                    if len(str(r[i])) > colWidths[i]:
                        colWidths[i] = len(str(r[i]))
                else:
                    colWidths += [len(r[i])]
                col = i + 2
                cell = ws.cell(rowNum, col)
                cell.value = r[i]
                if i == avgUnitsPos:
                    cell.number_format = "0.00"
                elif i == lineCountPos:
                    cell.number_format = "#,##0"
                elif i in dollarFields:
                    cell.number_format = "$#,##0.00"
                elif i in pctFields:
                    cell.number_format = "0.00%"
                # Customizations for header row
                if rowNum == 2:
                    cell.alignment = xlStyles.Alignment(horizontal="center")
                    cell.fill = xlStyles.PatternFill(
                        fgColor="e81c34", fill_type="solid"
                    )
                    cell.font = xlStyles.Font(color="FFFFFF", bold=True)
                # Global styles
                cell.border = thinBorder
        ws.auto_filter.ref = "B2:" + cellCoord
        for i, colWidth in enumerate(colWidths):
            ws.column_dimensions[get_column_letter(i + 2)].width = (colWidth + 6) * 1.2
    wb.save(xlout)
    print("done")

def getSql(queryName,tables):
    retSql = ""
    if queryName is "procPayerActivityClaims":
        retSql = """create table """ + tables["ppac"] + """ as (
            select remit_claim_service_id,
                    procedure_code,
                    payer_name,
                    toc_category,
                    payee_id,
                    payee_npi,
                    setting_name,
                    case when mod_codes is null or upper(mod_codes) = 'NULL' then 'NONE' else mod_codes end as mod_codes,
                    case when svc_prcsd_units is null then 0 else svc_prcsd_units end as svc_prcsd_units,
                    case when svc_allowed_amt is null then 0 else svc_allowed_amt end as svc_allowed_amt,
                    resA.Adjust_Reason_Code as Adjust_Reason_Code1,
                    resB.Adjust_Reason_Code as Adjust_Reason_Code2,
                    resC.Adjust_Reason_Code as Adjust_Reason_Code3,
                    resD.Adjust_Reason_Code as Adjust_Reason_Code4,
                    resE.Adjust_Reason_Code as Adjust_Reason_Code5,
                    resF.Adjust_Reason_Code as Adjust_Reason_Code6,
                    resA.Group_Code as Group_Code1,
                    resB.Group_Code as Group_Code2,
                    resC.Group_Code as Group_Code3,
                    resD.Group_Code as Group_Code4,
                    resE.Group_Code as Group_Code5,
                    resF.Group_Code as Group_Code6,
                    case when resA.Adjust_Amnt is null then '0' else resA.Adjust_Amnt end as adjust_Amnt1,
                    case when resB.Adjust_Amnt is null then '0' else resB.Adjust_Amnt end as adjust_Amnt2,
                    case when resC.Adjust_Amnt is null then '0' else resC.Adjust_Amnt end as adjust_Amnt3,
                    case when resD.Adjust_Amnt is null then '0' else resD.Adjust_Amnt end as adjust_Amnt4,
                    case when resE.Adjust_Amnt is null then '0' else resE.Adjust_Amnt end as adjust_Amnt5,
                    case when resF.Adjust_Amnt is null then '0' else resF.Adjust_Amnt end as adjust_Amnt6,
                    ppf.svc_sbmtd_chrg,
                    ppf.svc_line_item_paid_amt,
                    case when ppf.claim_patient_resp_amt is null then 0 else ppf.claim_patient_resp_amt end as claim_patient_resp_amt,
                    ppf.claim_status_code
            from """ + tables["claims"] + """ ppf
            left join """ + tables["adj"] + """ resA on ppf.svc_adjust_grp1 = resA.Adjust_Group_Id and ppf.job_id = resA.job_id
            left join """ + tables["adj"] + """ resB on ppf.svc_adjust_grp2 = resB.Adjust_Group_Id and ppf.job_id = resB.job_id
            left join """ + tables["adj"] + """ resC on ppf.svc_adjust_grp3 = resC.Adjust_Group_Id and ppf.job_id = resC.job_id
            left join """ + tables["adj"] + """ resD on ppf.svc_adjust_grp4 = resD.Adjust_Group_Id and ppf.job_id = resD.job_id
            left join """ + tables["adj"] + """ resE on ppf.svc_adjust_grp5 = resE.Adjust_Group_Id and ppf.job_id = resE.job_id
            left join """ + tables["adj"] + """ resF on ppf.svc_adjust_grp6 = resF.Adjust_Group_Id and ppf.job_id = resF.job_id
            where ppf.job_id = """ + tables["jobId"] + """)"""
    if queryName is "payPi":
        retSql = """create table """ + tables["pi"] + """ as (
        select pract.hms_piid,
                pract.first,
                pract.last,
                pract.address1,
                pract.address2,
                pract.city,
                pract.state,
                pract.zip
        from pri_plus_pract_profile pract where pract.job_id = '""" + tables["jobId"] + """'
    )"""
    if queryName is "payPo":
        retSql = """create table """ + tables["po"] + """ as (
        select org.hms_poid,
                org.org_name as orgname,
                org.address1,
                org.address2,
                org.city,
                org.state,
                org.zip
        from pri_plus_facility_profile org where org.job_id = '""" + tables["jobId"] + """'
    )"""
    if queryName is "procPayer":
        retSql = """create table """ + tables["pp"] + """ as (
        select pp.*,
                cpc.code_payer_count,
                cpsc.code_payer_spu_count
        from (select distinct procedure_code,payer_name,svc_prcsd_units from """ + tables["ppac"] + """) pp
        left join (
            select procedure_code,
                    payer_name,
                    count(*) as code_payer_count
            from """ + tables["ppac"] + """ group by procedure_code,payer_name) cpc
            on pp.procedure_code = cpc.procedure_code and pp.payer_name = cpc.payer_name
        left join (
            select procedure_code,
                    payer_name,
                    svc_prcsd_units,
                    count(*) as code_payer_spu_count
            from """ + tables["ppac"] + """ group by procedure_code,payer_name,svc_prcsd_units) cpsc
            on pp.procedure_code = cpsc.procedure_code and pp.payer_name = cpsc.payer_name and pp.svc_prcsd_units = cpsc.svc_prcsd_units
    )"""
    if queryName is "impProcPayer":
        retSql = """create table """ + tables["ipp"] + """ as (
        select distinct procedure_code,
                        payer_name from """ + tables["pp"]+ """
                        minus
        select procedure_code,
                payer_name
        from (
            select pp.*,
                    case
                      when pp.code_payer_spu_count/pp.code_payer_count >= .9
                        then 1
                      else 0
                    end as impute_units
            from """ + tables["pp"] + """ pp
            where svc_prcsd_units = 0
        )
        where impute_units = 1
    )"""
    if queryName is "cleanProcPayer":
        retSql = """create table """ + tables["cpp"] + """ as (
        select imp.procedure_code,
                imp.payer_name,
                imp.svc_prcsd_units,
                imp.imputation_value,
                case
                  when imp.procedure_code in (select procedure_code from """ + tables["ipp"] + """) and
                        imp.payer_name in (select payer_name from """ + tables["ipp"] + """)
                    then imp.svc_prcsd_units
                  when imp.imputation_value is null
                    then 1
                  else imp.imputation_value
                end as clean_units
        from (
            select pp.*,
                    iv.imputation_value
            from """ + tables["pp"] + """ pp
            left join (select procedure_code,
                                stats_mode(svc_prcsd_units) as imputation_value
                        from """ + tables["pp"] + """ where svc_prcsd_units > 0 and
                            procedure_code in (select procedure_code from """ + tables["ipp"] + """) and
                            payer_name in (select payer_name from """ + tables["ipp"] + """)
                            group by procedure_code) iv on pp.procedure_code = iv.procedure_code) imp
    )"""
    if queryName is "procPayerActivityClaimsClean":
        retSql = """create table """ + tables["ppacc"] + """ as (
            select ppac.*,
                   ppcpp.clean_units
            from (
                select * from """ + tables["ppac"] + """
                    minus
                select * from """ + tables["ppac"] + """ t
                        where t.svc_prcsd_units = 0 and
                        t.procedure_code in (select procedure_code from """ + tables["ipp"] + """) and
                        t.payer_name in (select payer_name from """ + tables["ipp"] + """)
                ) ppac
                left join """ + tables["cpp"] + """ ppcpp on
                    ppac.procedure_code = ppcpp.procedure_code and
                    ppac.payer_name = ppcpp.payer_name and
                    ppac.svc_prcsd_units = ppcpp.svc_prcsd_units
        )"""
    if queryName is "claimAdjAmts":
        retSql = """create table """ + tables["caa"] + """ as (
            select remit_claim_service_id, adjust_reason_code, sum(adjust_amnt) as adjust_amnt from (
                select remit_claim_service_id, adjust_reason_code1 as adjust_reason_code, adjust_amnt1 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code1 is not null
                    union
                select remit_claim_service_id, adjust_reason_code2 as adjust_reason_code, adjust_amnt2 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code2 is not null
                    union
                select remit_claim_service_id, adjust_reason_code3 as adjust_reason_code, adjust_amnt3 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code3 is not null
                    union
                select remit_claim_service_id, adjust_reason_code4 as adjust_reason_code, adjust_amnt4 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code4 is not null
                    union
                select remit_claim_service_id, adjust_reason_code5 as adjust_reason_code, adjust_amnt5 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code5 is not null
                    union
                select remit_claim_service_id, adjust_reason_code6 as adjust_reason_code, adjust_amnt6 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code6 is not null
                ) group by remit_claim_service_id, adjust_reason_code
        )"""
    if queryName is "patientRespTotal":
        retSql = """create table """ + tables["prt"] + """ as (
            select remit_claim_service_id,
                   sum(adjust_amnt) as patientRespTotal from (
            select remit_claim_service_id, adjust_reason_code, sum(adjust_amnt) as adjust_amnt from (
                select remit_claim_service_id, adjust_reason_code1 as adjust_reason_code, adjust_amnt1 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code1 in (1,2,3) or group_code1 = 'PR'
                    union
                select remit_claim_service_id, adjust_reason_code2 as adjust_reason_code, adjust_amnt2 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code2 in (1,2,3) or group_code2 = 'PR'
                    union
                select remit_claim_service_id, adjust_reason_code3 as adjust_reason_code, adjust_amnt3 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code3 in (1,2,3) or group_code3 = 'PR'
                    union
                select remit_claim_service_id, adjust_reason_code4 as adjust_reason_code, adjust_amnt4 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code4 in (1,2,3) or group_code4 = 'PR'
                    union
                select remit_claim_service_id, adjust_reason_code5 as adjust_reason_code, adjust_amnt5 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code5 in (1,2,3) or group_code5 = 'PR'
                    union
                select remit_claim_service_id, adjust_reason_code6 as adjust_reason_code, adjust_amnt6 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code6 in (1,2,3) or group_code6 = 'PR'
                ) group by remit_claim_service_id, adjust_reason_code) group by remit_claim_service_id
        )"""
    if queryName is "cleanCost":
        retSql = """create table """ + tables["cln"] + """ as (
            select remit_claim_service_id,
                   sum(adjust_amnt) as amnt_total from (
            select remit_claim_service_id, adjust_reason_code, sum(adjust_amnt) as adjust_amnt from (
                select remit_claim_service_id, adjust_reason_code1 as adjust_reason_code, adjust_amnt1 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code1 = """ + tables["ar"] + """
                    union
                select remit_claim_service_id, adjust_reason_code2 as adjust_reason_code, adjust_amnt2 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code2 = """ + tables["ar"] + """
                    union
                select remit_claim_service_id, adjust_reason_code3 as adjust_reason_code, adjust_amnt3 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code3 = """ + tables["ar"] + """
                    union
                select remit_claim_service_id, adjust_reason_code4 as adjust_reason_code, adjust_amnt4 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code4 = """ + tables["ar"] + """
                    union
                select remit_claim_service_id, adjust_reason_code5 as adjust_reason_code, adjust_amnt5 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code5 = """ + tables["ar"] + """
                    union
                select remit_claim_service_id, adjust_reason_code6 as adjust_reason_code, adjust_amnt6 as adjust_amnt from """ + tables["ppacc"] + """ where adjust_reason_code6 = """ + tables["ar"] + """
                ) group by remit_claim_service_id, adjust_reason_code) group by remit_claim_service_id
        )"""
    if queryName is "cleanVals1":
        retSql = """create table """ + tables["cv1"] + """ as (
            select prt.*,
                    case
                      when prt.patientresptotal > 0
                        then prt.patientresptotal
                      when prt.svc_allowed_amt > 0
                        then greatest(0,prt.svc_allowed_amt - prt.svc_line_item_paid_amt)
                      when prt.claim_patient_resp_amt > 0
                        then prt.claim_patient_resp_amt
                      else 0
                    end as clean_patientcost,
                    case
                      when prt.svc_sbmtd_chrg < 0
                        then null
                      else nvl(ded.amnt_total,0)
                    end as clean_deductible,
                    case
                      when prt.svc_sbmtd_chrg < 0
                        then null
                      else nvl(coins.amnt_total,0)
                    end as clean_coinsurance,
                    case
                      when prt.svc_sbmtd_chrg < 0
                        then null
                      else nvl(copay.amnt_total,0)
                    end as clean_copay,
                    case
                      when prt.svc_sbmtd_chrg < 0
                        then null
                      when prt.svc_allowed_amt > 0
                        then prt.svc_allowed_amt
                    else least(prt.svc_sbmtd_chrg,(prt.patientresptotal + prt.svc_line_item_paid_amt))
                    end as clean_allow
            from (
                select ppacc.*,
                        case
                          when svc_sbmtd_chrg < 0
                            then 0.00
                          else svc_sbmtd_chrg
                          end as clean_charge,
                        case
                          when svc_sbmtd_chrg < 0
                            then null
                          else svc_line_item_paid_amt
                          end as clean_paid,
                        case
                          when svc_sbmtd_chrg < 0
                            then null
                          else nvl(prt.patientresptotal,0)
                          end as patientRespTotal
                from """ + tables["ppacc"] + """ ppacc
                left join """ + tables["prt"] + """ prt on ppacc.remit_claim_service_id = prt.remit_claim_service_id) prt
                left join """ + tables["ded"] + """ ded on prt.remit_claim_service_id = ded.remit_claim_service_id
                left join """ + tables["coins"] + """ coins on prt.remit_claim_service_id = coins.remit_claim_service_id
                left join """ + tables["copay"] + """ copay on prt.remit_claim_service_id = copay.remit_claim_service_id
        )"""
    if queryName is "cleanVals2":
        retSql = """create table """ + tables["cv2"] + """ as (
            select cv2.*,
                    case
                      when cv2.clean_allow < cv2.clean_charge and cv2.clean_patientcost < cv2.clean_charge and cv2.paidBy <> 'Not Paid'
                        then 1
                      else 0
                      end as clean_discount,
                    cv2.clean_charge/cv2.clean_units as clean_chargeamtunit,
                    cv2.clean_allow/cv2.clean_units as clean_allowamtunit,
                    cv2.clean_paid/cv2.clean_units as clean_paidamtunit,
                    cv2.clean_patientcost/cv2.clean_units as clean_patientcostunit,
                    case
                      when ipp.procedure_code is null and ipp.payer_name is null and cv2.svc_prcsd_units = 0
                        then 1
                      else 0
                      end as imputed
                    from (
                        select cv1.*,
                                case
                                  when cv1.clean_patientcost = 0
                                    then null
                                  when cv1.clean_deductible > 0
                                    then 1
                                  else 0
                                  end as deductible_flag,
                                case
                                  when cv1.clean_patientcost = 0
                                    then null
                                  when cv1.clean_coinsurance > 0
                                    then 1
                                  else 0
                                  end as coinsurance_flag,
                                case
                                  when cv1.clean_patientcost = 0
                                    then null
                                  when cv1.clean_copay > 0
                                    then 1
                                  else 0
                                  end as copay_flag,
                                case
                                  when cv1.svc_sbmtd_chrg = 0
                                    then 'Unknown'
                                  when cv1.svc_sbmtd_chrg < 0 or cv1.claim_status_code = 22
                                    then 'Reversal'
                                  when cv1.claim_status_code = 4
                                    then 'Denied'
                                  when cv1.claim_status_code in (1,2,3,19,20,21)
                                    then case
                                      when cv1.clean_paid > 0 and cv1.clean_patientcost > 0
                                        then 'Shared Cost'
                                      when cv1.clean_paid = 0 and cv1.clean_patientcost > 0
                                        then 'Patient Only'
                                      when cv1.clean_paid > 0 and cv1.clean_patientcost = 0
                                        then 'Payer Only'
                                      when cv1.clean_paid = 0 and cv1.clean_patientcost = 0
                                        then 'Not Paid'
                                      else 'Unknown'
                                    end
                                  else 'Unknown'
                                  end as PaidBy
                        from """ + tables["cv1"] + """ cv1
                    ) cv2
                    left join """ + tables["ipp"] + """ ipp on cv2.procedure_code = ipp.procedure_code and cv2.payer_name = ipp.payer_name
        )"""
    if queryName is "claimAdjAmts2":
        retSql = """create table """ + tables["caa2"] + """ as (
            select caa.*,
                    case
                      when cv2.clean_units = 0 or cv2.clean_charge = 0
                        then 0
                      else case
                        when (caa.adjust_amnt/cv2.clean_units)/(cv2.clean_charge/cv2.clean_units) > 1
                          then 1
                        else (caa.adjust_amnt/cv2.clean_units)/(cv2.clean_charge/cv2.clean_units)
                      end
                    end as adjamtxunitperchrg
            from """ + tables["caa"] + """ caa
            left join """ + tables["cv2"] + """ cv2 on caa.remit_claim_service_id = cv2.remit_claim_service_id
        )"""
    if queryName is "cleanVals3":
        adjArr = tables["adjArr"]
        retSql = """create table """ + tables["cv3"] + """ as (
            select cv2.*"""
        for adjCode in adjArr:
            retSql += """,
                    case
                      when caa2""" + adjCode + """.adjamtxunitperchrg is null
                        then 0
                      else 1
                      end as code_""" + adjCode + """,
                      caa2""" + adjCode + """.adjamtxunitperchrg as adjamt""" + adjCode + """unitperchrg"""
        retSql += """
            from """ + tables["cv2"] + """ cv2"""
        for adjCode in adjArr:
            retSql += """
            left join (select * from """ + tables["caa2"] + " where adjust_reason_code = '""" + adjCode + """') caa2""" + adjCode + """ on cv2.remit_claim_service_id = caa2""" + adjCode + """.remit_claim_service_id"""
        retSql += """
        )"""
    if queryName is "acaReport":
        levels = tables["levels"]
        adjArr = tables["adjArr"]
        retSql = """select cv4.procedure_code,
                            cv4.payer_name,
                            cv4.line_of_business,
                            cv4.payee_id,
                            cv4.payee_npi,"""
        if levels:
            for level in levels:
                retSql += """
                            cv4.""" + level.lower() + ""","""
        retSql += """
                            cv4.paidBy as paid_by,
                            cv4.line_count,
                            cv4.avg_units,
                            case
                              when cv4.pct_units_imputed is not null
                                then cv4.pct_units_imputed
                              else null
                              end as pct_units_imputed,
                            
                            case
                              when cv4.chargeamtunit_avg is not null
                                then cv4.chargeamtunit_avg
                              else null
                              end as chargeamtunit_avg,
                            case
                              when cv4.chargeamtunit_median is not null
                                then cv4.chargeamtunit_median
                              else null
                              end as chargeamtunit_median,
                            case
                              when cv4.chargeamtunit_mode is not null
                                then cv4.chargeamtunit_mode
                              else null
                              end as chargeamtunit_mode,
                            case
                              when cv4.chargeamtunit_stddev is not null
                                then cv4.chargeamtunit_stddev
                              else null
                              end as chargeamtunit_stddev,
                            case
                              when cv4.chargeamtunit_min is not null
                                then cv4.chargeamtunit_min
                              else null
                              end as chargeamtunit_min,
                            case
                              when cv4.chargeamtunit_max is not null
                                then cv4.chargeamtunit_max
                              else null
                              end as chargeamtunit_max,
                            case
                              when cv4.allowedamtunit_avg is not null
                                then cv4.allowedamtunit_avg
                              else null
                              end as allowedamtunit_avg,
                            case
                              when cv4.allowedamtunit_median is not null
                                then cv4.allowedamtunit_median
                              else null
                              end as allowedamtunit_median,
                            case
                              when cv4.allowedamtunit_mode is not null
                                then cv4.allowedamtunit_mode
                              else null
                              end as allowedamtunit_mode,
                            case
                              when cv4.allowedamtunit_stddev is not null
                                then cv4.allowedamtunit_stddev
                              else null
                              end as allowedamtunit_stddev,
                            case
                              when cv4.allowedamtunit_min is not null
                                then cv4.allowedamtunit_min
                              else null
                              end as allowedamtunit_min,
                            case
                              when cv4.allowedamtunit_max is not null
                                then cv4.allowedamtunit_max
                              else null
                              end as allowedamtunit_max,
                            case
                              when cv4.paidamtunit_avg is not null
                                then cv4.paidamtunit_avg
                              else null
                              end as paidamtunit_avg,
                            case
                              when cv4.paidamtunit_median is not null
                                then cv4.paidamtunit_median
                              else null
                              end as paidamtunit_median,
                            case
                              when cv4.paidamtunit_mode is not null
                                then cv4.paidamtunit_mode
                              else null
                              end as paidamtunit_mode,
                            case
                              when cv4.paidamtunit_stddev is not null
                                then cv4.paidamtunit_stddev
                              else null
                              end as paidamtunit_stddev,
                            case
                              when cv4.paidamtunit_min is not null
                                then cv4.paidamtunit_min
                              else null
                              end as paidamtunit_min,
                            case
                              when cv4.paidamtunit_max is not null
                                then cv4.paidamtunit_max
                              else null
                              end as paidamtunit_max,
                            case
                              when cv4.patientcostunit_avg is not null
                                then cv4.patientcostunit_avg
                              else null
                              end as patientcostunit_avg,
                            case
                              when cv4.patientcostunit_median is not null
                                then cv4.patientcostunit_median
                              else null
                              end as patientcostunit_median,
                            case
                              when cv4.patientcostunit_mode is not null
                                then cv4.patientcostunit_mode
                              else null
                              end as patientcostunit_mode,
                            case
                              when cv4.patientcostunit_stddev is not null
                                then cv4.patientcostunit_stddev
                              else null
                              end as patientcostunit_stddev,
                            case
                              when cv4.patientcostunit_min is not null
                                then cv4.patientcostunit_min
                              else null
                              end as patientcostunit_min,
                            case
                              when cv4.patientcostunit_max is not null
                                then cv4.patientcostunit_max
                              else null
                              end as patientcostunit_max,
                            case
                              when cv4.pct_w_discount is not null
                                then cv4.pct_w_discount
                              else '0.00'
                              end as pct_w_discount,
                            case
                              when cv4.pct_w_deductible is not null
                                then cv4.pct_w_deductible
                              else '0.00'
                              end as pct_w_deductible,
                            case
                              when cv4.pct_w_copay is not null
                                then cv4.pct_w_copay
                              else '0.00'
                              end as pct_w_copay,
                            case
                              when cv4.pct_w_coinsurance is not null
                                then cv4.pct_w_coinsurance
                              else '0.00'
                              end as pct_w_coinsurance,"""
        for adjCode in adjArr:
            retSql += """
                            case
                              when cv4.adjAmt""" + adjCode + """UnitPerChrg is not null
                                then cv4.adjAmt""" + adjCode + """UnitPerChrg
                              else '0.00'
                              end as adjAmt""" + adjCode + """UnitPerChrg,
                            case
                              when cv4.pct_w_adjCode_""" + adjCode + """ is not null
                                then cv4.pct_w_adjCode_""" + adjCode + """
                              else '0.00'
                              end as pct_w_adjCode""" + adjCode + ""","""
        retSql += """
                            case
                              when cv4.payee_id like 'POZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_POID from """ + tables["po"] + """)
                                    then (select orgname from """ + tables["po"] + """ pol where cv4.payee_id = pol.hms_poid)
                                  else null
                                  end
                              when cv4.payee_id like 'PIZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_PIID from """ + tables["pi"] + """)
                                    then (select last from """ + tables["pi"] + """ pil where cv4.payee_id = pil.hms_piid) || ',' || (select first from """ + tables["pi"] + """ pil where cv4.payee_id = pil.hms_piid)
                                  else null
                                  end
                              else null
                              end as name,
                            case
                              when cv4.payee_id like 'POZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_POID from """ + tables["po"] + """)
                                    then (select address1 from """ + tables["po"] + """ pol where cv4.payee_id = pol.hms_poid)
                                  else null
                                  end
                              when cv4.payee_id like 'PIZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_PIID from """ + tables["pi"] + """)
                                    then (select address1 from """ + tables["pi"] + """ pil where cv4.payee_id = pil.hms_piid)
                                  else null
                                  end
                              else null
                              end as address1,
                            case
                              when cv4.payee_id like 'POZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_POID from """ + tables["po"] + """)
                                    then (select address2 from """ + tables["po"] + """ pol where cv4.payee_id = pol.hms_poid)
                                  else null
                                  end
                              when cv4.payee_id like 'PIZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_PIID from """ + tables["pi"] + """)
                                    then (select address2 from """ + tables["pi"] + """ pil where cv4.payee_id = pil.hms_piid)
                                  else null
                                  end
                                else null
                                end as address2,
                            case
                              when cv4.payee_id like 'POZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_POID from """ + tables["po"] + """)
                                    then (select city from """ + tables["po"] + """ pol where cv4.payee_id = pol.hms_poid)
                                  else null
                                  end
                              when cv4.payee_id like 'PIZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_PIID from """ + tables["pi"] + """)
                                    then (select city from """ + tables["pi"] + """ pil where cv4.payee_id = pil.hms_piid)
                                  else null
                                  end
                              else null
                              end as city,
                            case
                              when cv4.payee_id like 'POZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_POID from """ + tables["po"] + """)
                                    then (select state from """ + tables["po"] + """ pol where cv4.payee_id = pol.hms_poid)
                                  else null
                                  end
                              when cv4.payee_id like 'PIZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_PIID from """ + tables["pi"] + """)
                                    then (select state from """ + tables["pi"] + """ pil where cv4.payee_id = pil.hms_piid)
                                  else null
                                  end
                              else null
                              end as state,
                            case
                              when cv4.payee_id like 'POZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_POID from """ + tables["po"] + """)
                                    then (select zip from """ + tables["po"] + """ pol where cv4.payee_id = pol.hms_poid)
                                  else null
                                  end
                              when cv4.payee_id like 'PIZ%'
                                then case
                                  when cv4.payee_id in (select distinct HMS_PIID from """ + tables["pi"] + """)
                                    then (select zip from """ + tables["pi"] + """ pil where cv4.payee_id = pil.hms_piid)
                                  else null
                                  end
                              else null
                              end as zip
                            from
                            (
                                select cv3.procedure_code,
                                        cv3.payer_name,
                                        cv3.toc_category as line_of_business,
                                        cv3.payee_id,
                                        cv3.payee_npi,"""
        if levels:
            for level in levels:
                if level == "SETTING":
                    retSql += """
                                        cv3.setting_name as setting,"""
                else:
                    retSql += """
                                        cv3.""" + level.lower() + ""","""
        retSql += """
                                        cv3.paidby,
                                        count(*) as line_count,
                                        to_char(avg(cv3.clean_units),'fm999990.00') as avg_units,
                                        to_char(avg(cv3.imputed),'fm999990.00') as pct_units_imputed,
                                        to_char(avg(cv3.clean_chargeamtunit),'fm999990.00') as chargeamtunit_avg,
                                        to_char(median(cv3.clean_chargeamtunit),'fm999990.00') as chargeamtunit_median,
                                        to_char(stats_mode(cv3.clean_chargeamtunit),'fm999990.00') as chargeamtunit_mode,
                                        to_char(stddev(cv3.clean_chargeamtunit),'fm999990.00') as chargeamtunit_stddev,
                                        to_char(min(cv3.clean_chargeamtunit),'fm999990.00') as chargeamtunit_min,
                                        to_char(max(cv3.clean_chargeamtunit),'fm999990.00') as chargeamtunit_max,
                                        
                                        to_char(avg(cv3.clean_allowamtunit),'fm999990.00') as allowedamtunit_avg,
                                        to_char(median(cv3.clean_allowamtunit),'fm999990.00') as allowedamtunit_median,
                                        to_char(stats_mode(cv3.clean_allowamtunit),'fm999990.00') as allowedamtunit_mode,
                                        to_char(stddev(cv3.clean_allowamtunit),'fm999990.00') as allowedamtunit_stddev,
                                        to_char(min(cv3.clean_allowamtunit),'fm999990.00') as allowedamtunit_min,
                                        to_char(max(cv3.clean_allowamtunit),'fm999990.00') as allowedamtunit_max,
                                        
                                        to_char(avg(cv3.clean_paidamtunit),'fm999990.00') as paidamtunit_avg,
                                        to_char(median(cv3.clean_paidamtunit),'fm999990.00') as paidamtunit_median,
                                        to_char(stats_mode(cv3.clean_paidamtunit),'fm999990.00') as paidamtunit_mode,
                                        to_char(stddev(cv3.clean_paidamtunit),'fm999990.00') as paidamtunit_stddev,
                                        to_char(min(cv3.clean_paidamtunit),'fm999990.00') as paidamtunit_min,
                                        to_char(max(cv3.clean_paidamtunit),'fm999990.00') as paidamtunit_max,
                                        
                                        to_char(avg(cv3.clean_patientcostunit),'fm999990.00') as patientcostunit_avg,
                                        to_char(median(cv3.clean_patientcostunit),'fm999990.00') as patientcostunit_median,
                                        to_char(stats_mode(cv3.clean_patientcostunit),'fm999990.00') as patientcostunit_mode,
                                        to_char(stddev(cv3.clean_patientcostunit),'fm999990.00') as patientcostunit_stddev,
                                        to_char(min(cv3.clean_patientcostunit),'fm999990.00') as patientcostunit_min,
                                        to_char(max(cv3.clean_patientcostunit),'fm999990.00') as patientcostunit_max,
                                        
                                        to_char(avg(cv3.clean_discount),'fm999990.00') as pct_w_discount,
                                        to_char(avg(cv3.deductible_flag),'fm999990.00') as pct_w_deductible,
                                        to_char(avg(cv3.copay_flag),'fm999990.00') as pct_w_copay,
                                        to_char(avg(cv3.coinsurance_flag),'fm999990.00') as pct_w_coinsurance
                                        """
        for adjCode in adjArr:
            retSql += """,
                                        to_char(avg(cv3.adjamt""" + adjCode + """unitperchrg),'fm999990.00') as adjAmt""" + adjCode + """UnitPerChrg,
                                        case
                                          when cv3.paidby = 'Reversal'
                                            then null
                                          else to_char(avg(cv3.code_""" + adjCode + """),'fm999990.00')
                                        end as pct_w_adjCode_""" + adjCode
        retSql += """
                                from """ + tables["cv3"] + """ cv3
                                group by cv3.procedure_code,
                                            cv3.payer_name,
                                            cv3.toc_category,
                                            cv3.payee_id,
                                            cv3.payee_npi,"""
        if levels:
            for level in levels:
                if level == "SETTING":
                    retSql += """
                                        cv3.setting_name,"""
                else:
                    retSql += """
                                        cv3.""" + level.lower() + ""","""
        retSql += """
                                            cv3.paidby
                            ) cv4
                order by cv4.procedure_code,
                            cv4.payer_name,
                            cv4.line_of_business,
                            cv4.payee_id,
                            cv4.paidBy"""
        if levels:
            for level in levels:
                retSql += """,
                            cv4.""" + level.lower()
    return retSql

def styleRange(
    ws,
    cell_range,
    border=xlStyles.borders.Border(),
    fill=None,
    font=None,
    alignment=None,
):
    """
	Apply styles to a range of cells as if they were a single cell.
	
	:param ws:  Excel worksheet instance
	:param range: An excel range to style (e.g. A1:F20)
	:param border: An openpyxl Border
	:param fill: An openpyxl PatternFill or GradientFill
	:param font: An openpyxl Font object
	"""

    top = xlStyles.borders.Border(top=border.top)
    left = xlStyles.borders.Border(left=border.left)
    right = xlStyles.borders.Border(right=border.right)
    bottom = xlStyles.borders.Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

if __name__ == "__main__":
    startTime = time.time()
    main()
    endTime = time.time()
    hours, rem = divmod(endTime - startTime, 3600)
    minutes, seconds = divmod(rem, 60)
    print(
        "Program:\n\t"
        + __file__
        + "\nRun Time: {:0>2}:{:0>2}:{:05.2f}".format(int(hours), int(minutes), seconds)
    )
