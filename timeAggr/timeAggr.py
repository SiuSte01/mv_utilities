#!/opt/local/marketview/conda/1/envs/master/bin/python

from dateutil.parser import parse as dtParse
import logging
import argparse
import os,sys
import warnings
import json
import yaml
import pandas as pd
import numpy as np
from contextlib import closing
import itertools, collections
from tqdm import tqdm
import time
from datetime import datetime
import re
from mvUtil import miscfunctions as mf
import openpyxl
import pickle

warnings.simplefilter("ignore")
logger = logging.getLogger(__name__)
memoryHash = {
	"peopleHash": {},
	"errors": [],
	"custListFull": {},
	"custListMth": {},
	"timeStamp": datetime.now()
}
diffArr = []
empMap = {
	"BertEv01": "Bertel,Evan",
	"CharAl01": "Charlesworth,Allaire",
	"Nielja01": "Edgar-Nielsen,James",
	"EiseEm01": "Eisenberg,Emily",
	"EscoDa01": "Escott,David",
	"JoneMo01": "Jones,Molli",
	"KinzKa01": "Kinzer,Kathryn",
	"McClPa01": "McClelland,Patsy",
	"MeieCa01": "Meier,Carolyn",
	"OtteMa01": "Otterbein,Matthew",
	"PiatMa01": "Piatek,Mark",
	"QianJi01": "Qian,Jin",
	"SistJa01": "Sisti,Jason",
	"SiuSte01": "Siu,Stephen",
	"smitju02": "Smith,Justin",
	"StepRa01": "Stephano,Rachael",
	"TaylKa01": "Taylor,Karalee"
}

def main():
	#tqdmLoop()
	parser = argparse.ArgumentParser(description="timeSheet reader and reporter")
	parser.add_argument("-timeFrame", "--timeFrame")
	
	args = parser.parse_args()
	#print(args)
	#print (args.timeFrame)
	currTime = memoryHash["timeStamp"]
	currMth = str(currTime.month).zfill(2)
	currYear = str(currTime.year)
	trackRoot = "/vol/cs/clientprojects/MarketViewTracking"
	#print(trackRoot)
	clDir = trackRoot + "/timeAggr/customerLists"
	compDir = trackRoot + "/timeAggr/complianceReports"
	chlDir = trackRoot + "/timeAggr/changeLog"
	serDir = chlDir + "/serials"
	chgDir = chlDir + "/changeLogs"
	for d in [clDir,compDir,serDir,chgDir]:
		try:
			os.makedirs(d)
		except:
			pass
	if args.timeFrame is not None:
		dto = datetime.strptime(args.timeFrame,'%Y%m')
		currMth = str(dto.month).zfill(2)
		currYear = str(dto.year)
	
	for emp in os.listdir(trackRoot):
		#ignore files which are not subdirectories
		if not os.path.isdir(trackRoot + "/" + emp):
			continue
		#ignore timeAggr directory where time sheet states will be saved
		if emp == "timeAggr":
			continue
		print("Reading employee dir: " + emp + "...",end="")
		for timeSheet in os.listdir(trackRoot + "/" + emp):
			#only look at timesheet files
			if re.search("^(?!MVTimeTracker_\d{8}\.(csv|xlsx)).*$",timeSheet):
				continue
#			if emp not in ["SiuSte01"]:
#				continue
			#print("\t" + timeSheet)
			#
			if re.search("csv$",timeSheet):
				#print("csv")
				cifh = open(trackRoot + "/" + emp + "/" + timeSheet)
				header = re.sub("( *,|, *)",",",cifh.readline().strip('\ufeff').rstrip())
				headers = mf.getHeaders(convertToTab(header))
				#print(headers)
				parseIterable(cifh.readlines(),headers,'csv',emp,timeSheet)
			else:
				#print("xlsx")
				wb = openpyxl.load_workbook(trackRoot + "/" + emp + "/" + timeSheet)
				ws = wb.active
				headerArr = []
				#print(ws.max_row)
				for hdr in ws[1]:
					cleanHdr = hdr.value.strip()
					headerArr.append(cleanHdr)
				headers = mf.getHeaders(convertToTab(','.join(headerArr)))
				#print(headers)
				parseIterable(ws.iter_rows(min_row=2),headers,'xlsx',emp,timeSheet)
		print("done")
	print("")
	prevMemoryHash = {}
	serFile = serDir + "/../latestSerial.txt"
	currMemTime = ''.join([str(memoryHash["timeStamp"].year),str(memoryHash["timeStamp"].month),str(memoryHash["timeStamp"].day)])
	if os.path.isfile(serFile):
		#load previous memoryHash and compare entry differences
		print("Previous serial file found. Checking for changes...",end="")
		prevMemoryHash = pickle.load(open(serFile,"rb"))
		#print(yaml.dump(prevMemoryHash))
		#print("\n\nthis happens\n\n")
		#print(yaml.dump(memoryHash))
		#exit()
		for emp in memoryHash["peopleHash"]:
			#new employee encountered
			if not emp in prevMemoryHash["peopleHash"]:
				diffArr.append("New employee encountered: " + emp)
				continue
			#employee found in prevMemoryHash, compare entries
			for entry in memoryHash["peopleHash"][emp]["workEntries"]:
				if not entry in prevMemoryHash["peopleHash"][emp]["workEntries"]:
					diffArr.append("New entry encountered: " + emp + " - " + entry)
		for emp in prevMemoryHash["peopleHash"]:
			#employee missing from new run
			if not emp in memoryHash["peopleHash"]:
				diffArr.append("Previously listed employee missing: " + emp)
				continue
			#employee found in memoryHash, compare entries
			for entry in prevMemoryHash["peopleHash"][emp]["workEntries"]:
				if not entry in memoryHash["peopleHash"][emp]["workEntries"]:
					diffArr.append("Previously listed entry missing: " + emp + " - " + entry)
					continue
				#entry found, check to see if any data has changed
				oldEnt = prevMemoryHash["peopleHash"][emp]["workEntries"][entry]
				newEnt = memoryHash["peopleHash"][emp]["workEntries"][entry]
				entryChanges = []
				if oldEnt["hours"] != newEnt["hours"]:
					entryChanges.append("Hours: " + str(oldEnt["hours"]) + "->" + str(newEnt["hours"]))
				if oldEnt["deliverableComplete"] != newEnt["deliverableComplete"]:
					entryChanges.append("Deliverable Complete: " + oldEnt["deliverableComplete"] + "->" + newEnt["deliverableComplete"])
				if oldEnt["nonStandardProc"] != newEnt["nonStandardProc"]:
					entryChanges.append("Non-Standard Process: " + oldEnt["nonStandardProc"] + "->" + newEnt["nonStandardProc"])
				if oldEnt["workTypes"] != newEnt["workTypes"]:
					entryChanges.append("WorkTypes: " + ','.join(sorted(oldEnt["workTypes"])) + "->" + ','.join(sorted(newEnt["workTypes"])))
				if entryChanges:
					diffArr.append("Entry values have changed: " + emp + " - " + entry + "\n\t" + "\n\t".join(entryChanges))
		#there are differences, create a changelog for timestamp from prevMemHash and memHash
		if diffArr:
			print("changes found!\n\tPlease check changeLog.txt file in " + chlDir + " to see differences\n")
			cofh = open(chgDir + "/../changeLog.txt","w")
			prevMemTime = ''.join([str(prevMemoryHash["timeStamp"].year),str(prevMemoryHash["timeStamp"].month),str(prevMemoryHash["timeStamp"].day)])
			cofh.write("Differences between " + prevMemTime + " and " + currMemTime + "\n")
			for x in diffArr:
				cofh.write(x + "\n")
			cofh.close()
			timeStampedChangeLog = currMemTime + "Log.txt"
			os.system("cp " + chgDir + "/../changeLog.txt " + chgDir + "/" + timeStampedChangeLog)
		else:
			print("None found\n")
	#print(yaml.dump(diffArr))
	#print(prevMemoryHash["timeStamp"])
	#print(memoryHash["timeStamp"])
	#exit()
	
	print("Updating latestSerial.txt file...",end="")
	sofh = open(serDir + "/../latestSerial.txt","wb")
	pickle.dump(memoryHash,sofh)
	sofh.close()
	timeStampedSerial = currMemTime + "Serial.txt"
	os.system("cp " + serDir + "/../latestSerial.txt " + serDir + "/" + timeStampedSerial)
	print("done\n")
	#exit()
	#print(yaml.dump(memoryHash))
	#print(yaml.dump(memoryHash["custListFull"]))
	#print(yaml.dump(memoryHash["custListMth"]))
	#print(len(memoryHash["custListFull"]))
	#print(len(memoryHash["custListMth"]))
	#write compliance data
	print("Generating latest complianceReport file...",end="")
	crofh = open(''.join([compDir,"/complianceReport",currYear,currMth,".txt"]),"w")
	crofh.write("#############EMPLOYEE METADATA#############\n\n")
	for employee in sorted(memoryHash["peopleHash"]):
		crofh.write(employee + "\n")
		lastEntry = memoryHash["peopleHash"][employee]["metaData"]["lastEntryDate"]
		lastSheet = memoryHash["peopleHash"][employee]["metaData"]["lastTimeSheetDate"]
		crofh.write("\tLast Dated Entry: " + lastEntry + "\n")
		crofh.write("\tLast Dated Time Sheet: " + lastSheet + "\n")
	crofh.write("\n")
	crofh.write("#############CURRENT ERRORS#############\n\n")
	for x in memoryHash["errors"]:
		crofh.write(x)
	crofh.close()
	print("done\n")
	#write customer lists
	print("Updating customerList files...",end="")
	clfofh = open(clDir + "/customerListFull.tab","w")
	clmofh = open(''.join([clDir,"/customerList",currYear,currMth,".tab"]),"w")
	for ofh in [clfofh,clmofh]:
		ofh.write("CLIENT\tEMPLOYEES\tTOTAL_HOURS\n")
	for cl in ["custListFull","custListMth"]:
		for cust in sorted(memoryHash[cl],key=lambda s: s.lower()):
			ofh = clfofh if cl == "custListFull" else clmofh
			emps = ":".join(sorted(memoryHash[cl][cust]["employees"]))
			totalHours = str(memoryHash[cl][cust]["totalHours"])
			ofh.write("\t".join([cust,emps,totalHours]) + "\n")
		ofh.close()
	print("done\n")

def convertToTab(line):
	#this code will take in a string of characters and parse through it and convert the appropriate commas to
	#tabs and leave the literal commas alone
	#it will return the fixed string for the code that calls this function to use
	sep = '\t'
	quoted = False
	outList = []
	for char in line:
		if char == "\"":
			quoted = not quoted
		elif char == ",":
			if quoted:
				outList.append(char)
			else:
				outList.append(sep)
		else:
			outList.append(char)
	return ''.join(outList)

def parseIterable(itr,headers,itrType,emp,timeSheet):
	tsDate = re.sub("(^.*_|\..*$)","",timeSheet)
	mappedUser = empMap[emp] if emp in empMap else ""
	#print('-'.join([itrType,emp,mappedUser,timeSheet,tsDate]))
	if mappedUser:
		if not mappedUser in memoryHash["peopleHash"].keys():
			memoryHash["peopleHash"][mappedUser] = {
				"workEntries": {},
				"metaData": {
					"lastEntryDate": "19700101",
					"lastTimeSheetDate": tsDate
				}
			}
		else:
			if tsDate > memoryHash["peopleHash"][mappedUser]["metaData"]["lastTimeSheetDate"]:
				memoryHash["peopleHash"][mappedUser]["metaData"]["lastTimeSheetDate"] = tsDate
	recordCounter = 0
	#print(headers)
	for row in itr:
		recordCounter += 1
		entry = []
		if itrType == "csv":
			line = convertToTab(row.rstrip())
			#print(line)
			entry = line.split('\t')
			entry = [x.strip(' ') for x in entry]
			entry = [x.strip('"') for x in entry]
		else:
			for cell in row:
				if cell.value is None:
					entry.append("")
				elif cell.column == headers["Date"] + 1:
					m = str(cell.value.month)
					d = str(cell.value.day)
					y = str(cell.value.year)
					entry.append('/'.join([m,d,y]))
				else:
					entry.append(str(cell.value))
					#print(f"col {cell.col_idx} : {cell.value}")
		#print(entry)
		userRaw = entry[headers['User']]
		date = mf.normalizeDate(entry[headers['Date']],yyyymmdd='y')
		cust = entry[headers['Customer']]
		task = entry[headers['Task']]
		hours = entry[headers['Hours']]
		#skip line if all fields are empty (no need to check to user)
		if all(field == "" for field in entry[1:]):
			continue
		#there are populated fields but the 4 required fields are empty, mark row as error row
		elif any(field == "" for field in [userRaw,date,cust,task]):
			memoryHash["errors"].append(emp + "/" + timeSheet + " - Bad row on line: " + str(recordCounter) + "\n" +
				"\t" + ','.join(entry))
			continue
		#user does not have a comma in it, implying it is not of form last,first
		elif not re.search(",",userRaw):
			memoryHash["errors"].append(emp + "/" + timeSheet + " - User does not follow 'Last,First' naming convention on line: " + str(recordCounter) + "\n" +
				"\t" + ','.join(entry))
			continue
		#initial checks passed, normalize username
		userArr = userRaw.split(',')
		userArr = [x.strip(' ') for x in userArr]
		userArr = [x[0].upper() + x[1:] for x in userArr]
		user = ','.join(userArr)
		
		dateMth = date[4:6]
		dateYr = date[0:4]
		currMth = str(memoryHash["timeStamp"].month).zfill(2)
		currYear = str(memoryHash["timeStamp"].year)
		#print(date)
		#print(dateMth)
		#print(currMth)
		#lineTracker
		if not cust in memoryHash["custListFull"].keys():
			memoryHash["custListFull"][cust] = {
				"totalHours": float(0),
				"employees": set()
			}
		memoryHash["custListFull"][cust]["totalHours"] += float(hours)
		memoryHash["custListFull"][cust]["employees"].update([user])
		if int(dateMth) == int(currMth) and int(dateYr) == int(currYear):
			if not cust in memoryHash["custListMth"].keys():
				memoryHash["custListMth"][cust] = {
					"totalHours": float(0),
					"employees": set()
				}
			memoryHash["custListMth"][cust]["totalHours"] += float(hours)
			memoryHash["custListMth"][cust]["employees"].update([user])
		#print(timeSheet)
		#print("\t" + tsDate)
		#print("\t" + '-'.join(entry))
		deliv = entry[headers['Deliverable Complete']]
		nonStandardProc = entry[headers['Non-Standard Process']]
		workTypes = ['PxDx','INA','PRI','Trend','EA','ActRpt',
		'DistRpt','PayMix','SOC','SL','PAC','Wrkbk','UI','Other']
		entryKey = re.sub(" ","",''.join([date,cust,task]))
		#user has not been defined. create first entry
		#this should ONLY happen if encountering a User directory we don't have a mapping for
		if not user in memoryHash["peopleHash"].keys():
			memoryHash["peopleHash"][user] = {
				"workEntries": {
					entryKey: {
						"date": date,
						"customer": cust,
						"task": task,
						"hours": float(hours),
						"deliverableComplete": "Y" if deliv.upper() == "Y" else "N",
						"nonStandardProc": "Y" if nonStandardProc.upper() == "Y" else "N",
						"workTypes": set(),
						"encounters": 1
					}
				},
				"metaData": {
					"lastEntryDate": date,
					"lastTimeSheetDate": tsDate
				}
			}
			for workType in workTypes:
				if entry[headers[workType]].upper() != "":
					memoryHash["peopleHash"][user]["workEntries"][entryKey]["workTypes"].update([workType])
		#user exists, but workEntry does not, add a new entry
		if not entryKey in memoryHash["peopleHash"][user]["workEntries"].keys():
			memoryHash["peopleHash"][user]["workEntries"][entryKey] = {
				"date": date,
				"customer": cust,
				"task": task,
				"hours": float(hours),
				"deliverableComplete": "Y" if deliv.upper() == "Y" else "N",
				"nonStandardProc": "Y" if nonStandardProc.upper() == "Y" else "N",
				"workTypes": set(),
				"encounters": 1
			}
			if date > memoryHash["peopleHash"][user]["metaData"]["lastEntryDate"]:
				memoryHash["peopleHash"][user]["metaData"]["lastEntryDate"] = date
			if tsDate > memoryHash["peopleHash"][user]["metaData"]["lastTimeSheetDate"]:
				memoryHash["peopleHash"][user]["metaData"]["lastTimeSheetDate"] = tsDate
			for workType in workTypes:
				if entry[headers[workType]].upper() != "":
					memoryHash["peopleHash"][user]["workEntries"][entryKey]["workTypes"].update([workType])
		#user exists and workEntry also exists. update existing entry
		else:
			memoryHash["peopleHash"][user]["workEntries"][entryKey]["hours"] += float(hours)
			if deliv.upper() == "Y":
				memoryHash["peopleHash"][user]["workEntries"][entryKey]["deliverableComplete"] = "Y"
			if nonStandardProc.upper() == "Y":
				memoryHash["peopleHash"][user]["workEntries"][entryKey]["nonStandardProc"] = "Y"
			for workType in workTypes:
				if entry[headers[workType]].upper() != "":
					memoryHash["peopleHash"][user]["workEntries"][entryKey]["workTypes"].update([workType])
			memoryHash["peopleHash"][user]["workEntries"][entryKey]["encounters"] += 1
		#print("\t\t" + ':'.join([user,date,cust,task,hours,deliv,nonStandardProc,','.join(workTypes)]))

def tqdmLoop():
	for i in tqdm(range(0,100),desc="tqdm sample"):
		sleep(.5)

if __name__ == '__main__':
	startTime = time.time()
	main()
	endTime = time.time()
	hours, rem = divmod(endTime-startTime, 3600)
	minutes, seconds = divmod(rem, 60)
	print("Program:\n\t" + __file__ + "\nRun Time: {:0>2}:{:0>2}:{:05.2f}".format(int(hours),int(minutes),seconds))
